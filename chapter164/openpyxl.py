import openpyxl as opx 
from openpyxl import Workbook 

path = "C:/Users/MyHP/Desktop/PYTHONbook/chapter164/datafile.xlsx"
workbook = opx.load_workbook("C:/Users/MyHP/Desktop/PYTHONbook/chapter164/datafile.xlsx")
first_sheet = workbook.worksheets[0]

#If you want to specify the name of an available sheets, you can use workbook.get_sheet_names().
sheet = workbook.get_sheet_by_name('Sheet Name')

"""Creating a new Workbook in memory"""

wb = opx.Workbook()
ws = wb.create_sheet('Sheet Name', 0)
ws.sheet_properties.tabColor = 'FFC0CB'
wb.save('data.xlsx') 